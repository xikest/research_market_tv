import os
import pickle
from openpyxl import Workbook, load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from datetime import datetime
import pandas as pd

class FileManager:
    def __init__(self):
        pass

    @staticmethod
    def make_dir(folder_path):
        relative_path = folder_path.split("/")
        current_path = os.getcwd()
        for folder_name in relative_path:
            current_path = os.path.join(current_path, folder_name)
            if not os.path.exists(current_path):
                os.makedirs(current_path)
                print(f"폴더 '{folder_name}'가 생성되었습니다.")

    @staticmethod
    def save_file(data, filename="backup"):
        with open(f"{filename}.pickle", "wb") as file:
            pickle.dump(data, file)
        return data

    @staticmethod
    def load_file(filename="backup"):
        with open(f"{filename}.pickle", 'rb') as file:
            data = pickle.load(file)
        return data

    @staticmethod
    def get_datetime_info(include_time=True):
        format_str = "%Y-%m-%d_%H-%M-%S" if include_time else "%Y-%m-%d"
        return datetime.now().strftime(format_str)

    @staticmethod
    def get_name_from_url(url):
        return url.rsplit('/', 1)[-1]

    @staticmethod
    def dict_to_excel(dict_data, file_name="dictToExcel", sheet_name="sheet1", orient_idx=True):
        orient = 'columns'
        if orient_idx: orient = 'index'

        df = pd.DataFrame.from_dict(dict_data, orient=orient).reset_index().rename(columns={'index': 'model'})

        try:
            wb = load_workbook(filename=f"{file_name}.xlsx")
            ws = wb.create_sheet(title=sheet_name)
        except FileNotFoundError:
            wb = Workbook()
            ws = wb.active
            ws.title = sheet_name

        for row in dataframe_to_rows(df, index=False, header=True):
            ws.append(row)

        wb.save(f"{file_name}.xlsx")
        print(f"데이터가 {file_name}.xlsx 파일의 {sheet_name} 시트에 저장되었습니다.")
        return dict_data
