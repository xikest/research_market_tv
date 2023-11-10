import os
import time
import pickle
from datetime import datetime
import random
from openpyxl import Workbook, load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
import pandas as pd

def generate_random_number(start:int=3, end:int=5):
    return random.randint(start, end)


def makeDir(folder_path):
    # 현재 작업 디렉토리 기준으로 상대 경로 설정
    relative_path = folder_path.split("/")
    # 폴더 생성을 위해 순차적으로 경로 조합
    current_path = os.getcwd()
    for folder_name in relative_path:
        current_path = os.path.join(current_path, folder_name)
        # 폴더가 존재하지 않을 경우에만 폴더 생성
        if not os.path.exists(current_path):
            os.makedirs(current_path)
            print(f"폴더 '{folder_name}'가 생성되었습니다.")
        # else:
            # print(f"폴더 '{folder_name}'가 이미 존재합니다.")


def waitingPage(sec:int=5):
    time.sleep(sec) # 웹 페이지 대기
    return None

def backUp(data, filename:str="backup"):
    # 데이터 저장
    with open(f"{filename}.pickle", "wb") as file:
        pickle.dump(data, file)
    # 데이터 불러오기
    with open(f"{filename}.pickle", "rb") as file:
        data = pickle.load(file)
    return data


def load_pickle_file(filename:str="backup"):
    with open(f"{filename}.pickle", 'rb') as file:
        data = pickle.load(file)
    return data

def get_time():# 현재 시간 얻기
    return datetime.now().strftime("%Y-%m-%d_%H-%M-%S")

def get_today():
    return datetime.today().strftime("%Y-%m-%d")

def getNamefromURL(url):
    """
    마지막 / 이후 문자만 가져 옴
    """
    return url.rsplit('/', 1)[-1]

def dictToexcel(dictData, fileName="dictToExcel", sheetName="sheet1", orient_idx=True):
    if orient_idx:
        orient = 'index'
    else:
        orient = 'columns'

    df = pd.DataFrame.from_dict(dictData, orient=orient).reset_index().rename(columns={'index': 'model'})

    try:
        wb = load_workbook(filename=f"{fileName}.xlsx")
        ws = wb.create_sheet(title=sheetName)
    except FileNotFoundError:
        wb = Workbook()
        ws = wb.active
        ws.title = sheetName

    for row in dataframe_to_rows(df, index=False, header=True):  # 수정된 부분
        ws.append(row)

    wb.save(f"{fileName}.xlsx")
    print(f"데이터가 {fileName}.xlsx 파일의 {sheetName} 시트에 저장되었습니다.")
    return dictData
